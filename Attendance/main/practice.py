import cv2
import numpy as np
import pickle
import openpyxl
import button
import datetime
from openpyxl.styles import PatternFill
import os

button = [20,60,10,350]

def process_click(event, x, y,flags, params):
    if event == cv2.EVENT_LBUTTONDOWN and y > button[0] and y < button[1] and x > button[2] and x < button[3]: # condition for mouse position in between button borders
        # print('Clicked on Button!')
        exec(open('../button/add_face.py').read())
        print("PROCESSING DATA...")
        exec(open('../main/practice-train.py').read())
        exec(open('../main/practice.py').read())

date = datetime.datetime.now()
present_date = date.date()

hour = date.hour
minS = date.minute
sec = date.second
# print('hr:', hour, 'mn:',  minS, 'sc:', sec)
tm = 'Time = ' + str(hour) + '-' + str(minS) + '-' + str(sec)

isFile = os.path.isfile('../Attendance/attendance_' + str(present_date) + '.xlsx')

if (isFile):
    wb = openpyxl.load_workbook('../Attendance/attendance_' + str(present_date) + '.xlsx')
    print('Loading file...')
else:
    wb = openpyxl.Workbook()
    print('Creating file...')

sHeet = wb.active
sHeet['A1'] = 'Attendance: ' + str(present_date)
sheet = wb.create_sheet('Mysheet')
sheet.title = tm

att = []

face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read("trainner.yml")

labels = {"person_name": 1}
with open('labels.pickle', 'rb') as f:
    og_labels = pickle.load(f)
    labels = {v:k for k,v in og_labels.items()}
    totCnt = max(labels, key=labels.get)
    # print(totCnt)
    sheet['A1'] = 'Student name'
    sheet['B1'] = present_date
    num = 2
    for i in range(0, totCnt + 1):
        # print(labels[i])
        sheet['A' + str(num)] = labels[i]
        sheet['B' + str(num)] = 'A'
        sheet['B' + str(num)].fill = PatternFill(start_color='ffcccb', fill_type='solid')
        num += 1


cap = cv2.VideoCapture(0)
while True:
    ret, image = cap.read(0)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.5, minNeighbors=5)
    for (x, y, w, h) in faces:
        cv2.rectangle(image, (x, y), (x + w, y + h), (255, 0, 0), 2)
        roi_gray = gray[y:y + h, x:x + w]
        roi_color = image[y:y + h, x:x + w]
        id_, conf = recognizer.predict(roi_gray)  # id_ returns labels back from practice-train.py
        if conf >= 45 and conf <= 85:
            # print(id_)
            # print(labels[id_])
            name = labels[id_]
            cv2.putText(image, name.upper(), (x, y), cv2.FONT_HERSHEY_PLAIN, 2, (255, 0, 0), 2, cv2.LINE_AA)
            att.append(id_)
        else:
            cv2.putText(image, 'Unknown', (x, y), cv2.FONT_HERSHEY_PLAIN, 2, (255, 0, 0), 2, cv2.LINE_AA)

        cv2.imwrite('7.png', roi_color)  # it save the b/w image of only face which is detected
    # cv2.namedWindow('image', cv2.WND_PROP_FULLSCREEN)
    # cv2.setWindowProperty('image', cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)


    for i in range(0, totCnt + 1):
        if att.count(i) >= 50 and att.count(i) <= 70:
            # print(labels[i] + " : present")
            # print("Count : ", att.count(i))
            sheet['B' + str(i + 2)] = 'P'
            sheet['B' + str(i + 2)].fill = PatternFill(start_color='90ee90', fill_type='solid')
            cv2.putText(image, labels[i].upper() + " MARKED PRESENT", (20, 40), cv2.FONT_HERSHEY_PLAIN, 2, (0, 255, 0), 2, cv2.LINE_AA)

    cv2.namedWindow('Control')
    cv2.setMouseCallback('Control', process_click)

    control_image = np.zeros((80, 360), np.uint8)  # for creating background as black frame
    control_image[button[0]:button[1], button[2]:button[3]] = 180
    cv2.putText(control_image, 'ADD STUDENT FACE', (20, 50), cv2.FONT_HERSHEY_PLAIN, 2, (0), 3)  # text on button
    cv2.imshow('Control', control_image)

    cv2.imshow("Attendance", image)

    k = cv2.waitKey(1)
    if (k == ord('q')):
        break


wb.save('../attendance/Attendance_' + str(present_date) + '.xlsx')
cap.release()
cv2.destroyAllWindows()